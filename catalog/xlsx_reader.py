"""Thin wrapper around openpyxl: dump a worksheet as a list of rows of plain
Python values (str / int / float / None), left-stripped of fully-empty
trailing rows/columns. No interpretation of content happens here — that is
the job of catalog/parsing.py.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl


def read_sheet_rows(path: Path, sheet_name: str) -> list[list[object]]:
    """Return every row of `sheet_name` in `path` as a list of cell values.

    Uses read_only mode for speed; values come back as openpyxl already
    infers them (numbers as int/float, text as str, blanks as None).
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        rows: list[list[object]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        return rows
    finally:
        wb.close()


def sheet_names(path: Path) -> list[str]:
    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()
